"""
Data export endpoints: CSV and Excel.
"""
import io
import csv
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Server, ServerMetric

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/csv")
def export_csv(db: Session = Depends(get_db)):
    """Export server logs as CSV."""
    servers = db.query(Server).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Timestamp", "Server ID", "Server Name", "IP Address",
        "Airport Code", "Location", "Status", "Health Score",
        "Latency (ms)", "Packet Loss (%)", "Uptime (%)",
        "CPU Usage (%)", "Memory Usage (%)", "Response Time (ms)"
    ])

    for server in servers:
        metrics = (
            db.query(ServerMetric)
            .filter(ServerMetric.server_id == server.id)
            .order_by(ServerMetric.timestamp.desc())
            .limit(100)
            .all()
        )
        for m in metrics:
            writer.writerow([
                str(m.timestamp),
                server.id,
                server.name,
                server.ip_address,
                server.airport_code,
                server.location_name,
                server.status,
                server.health_score,
                m.latency_ms,
                m.packet_loss,
                m.uptime_percent,
                m.cpu_usage,
                m.memory_usage,
                m.response_time,
            ])

    output.seek(0)
    filename = f"server_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/excel")
def export_excel(db: Session = Depends(get_db)):
    """Export server logs as Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Server Logs"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="001F6D", end_color="001F6D", fill_type="solid")

    headers = [
        "Timestamp", "Server ID", "Server Name", "IP Address",
        "Airport Code", "Location", "Status", "Health Score",
        "Latency (ms)", "Packet Loss (%)", "Uptime (%)",
        "CPU Usage (%)", "Memory Usage (%)", "Response Time (ms)"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    servers = db.query(Server).all()
    row_num = 2
    for server in servers:
        metrics = (
            db.query(ServerMetric)
            .filter(ServerMetric.server_id == server.id)
            .order_by(ServerMetric.timestamp.desc())
            .limit(100)
            .all()
        )
        for m in metrics:
            data = [
                str(m.timestamp), server.id, server.name, server.ip_address,
                server.airport_code, server.location_name, server.status,
                server.health_score, m.latency_ms, m.packet_loss,
                m.uptime_percent, m.cpu_usage, m.memory_usage, m.response_time,
            ]
            for col, val in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=val)
            row_num += 1

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"server_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
